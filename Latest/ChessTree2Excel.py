from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from datetime import date, datetime, timedelta

class searchTree():
    """ very basic class to hold search and save it to excel. """

    def __init__(self, maxDepth) -> None:
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "SearchTree"
        self.currentRow = 1
        self.maxDepth = maxDepth

    def addSheet(self, name):
        """
        Can be used to create some new worksheets. Currently not used.
        in a future version to put all moves in same workbook. 1 move per sheet.
        """

        self.wb.create_sheet(name)

    def insertData(self, depth, data, eval, topMove):
        """ This insert one row of data in our Excel sheet. """

        # Column 1 is the current best move. Starts empty and gets updated as we search
        self.ws.cell(row=self.currentRow, column=1, value = topMove)

        # This writes the move sequence in the right column based on the # of moves in data
        self.ws.cell(row=self.currentRow, column= 1 + len(data), value = str(data[-1]))
        self.ws.cell(row=self.currentRow, column= 2 + self.maxDepth, value = str([str(m) for m in data]))
        self.ws.cell(row=self.currentRow, column= 3 + self.maxDepth, value = round(eval,2))

        self.currentRow += 1
    
    def format(self):

        self.ws.sheet_view.zoomScale = 200

        # Inserting the header row
        self.ws.insert_rows(1)

        # labels each column
        for y in range(1,self.maxDepth+1):
            self.ws.cell(row=1, column=1, value = "CurrentBest")
            self.ws.cell(row=1, column=1 + y, value = f"Level {y}")
            self.ws.cell(row=1, column=2 + self.maxDepth, value = "Sequence")
            self.ws.cell(row=1, column=3 + self.maxDepth, value = "Eval")

        # Formatting the header row
        for y in range(1 , self.maxDepth + 4):
            myCell = self.ws.cell(row=1, column=y)
            myCell.font = Font(bold=True)
            myCell.alignment = Alignment(horizontal="center", vertical="center")
            myCell.fill = PatternFill("solid", fgColor="DDDDDD")

        # Resize columns ['A', 'B', ... ] for the length of headerrow. 
        for i in range (0,self.maxDepth+3):
            self.ws.column_dimensions[chr(65+i)].width = 10
        self.ws.column_dimensions[chr(65+self.maxDepth + 1)].width = 30

    def saveToExcel(self, name):
        """ Saving our workbook to disk name with date/time + move number.xls """
        
        self.format()

        filename = datetime.now().strftime("%y%m%d-%H%M%S") + name
        self.wb.save(filename + '.xlsx')
